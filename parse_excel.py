#Imports
from openpyxl import load_workbook

"""
error file not found

input: filename

Description:
Given a filename, prints an error message that the given file was not found
Then exits the program
"""
def error_file_not_found(filename):
    print("Error: The file", filename, "was not found")
    exit()

"""
load excel spreadsheet partial

input: (string filename, int column)
output: a list containing all data entries for a given column

Description: 
Given the path for an excel spreadsheet and a column number,
load all entries from that column and return them as a list.

Note: This works with all spreadsheets but doesn't strictly define datatypes
"""
def load_excel_spreadsheet_partial(filename, column):
	data = []

	try:
		workbook = load_workbook(filename) 		#Load .xlsx file
	except:
	    error_file_not_found(filename)
	
	sheet = workbook.active				#Current excel sheet (we are only using one)
	rows = sheet.rows				#Rows from excel sheet
	max_row = sheet.max_row				#Gets count of max row
	
	"""
	Why start the loop at 2:
		We do not need the first row which contains the headers i.e. subject, course #, etc.
		Additionally, iterating through an excel spreadsheet starts at 1 not 0.		
		
	Why end the loop at max_row + 1:
		Since the iteration starts at one, we need to add +1 at the end to get every row
	"""
	for i in range(2, max_row + 1):
		
		entry = sheet.cell(row=i, column=column) #single entry
		data.append(entry.value)
	
	return data

"""
load schedule hardcoded

input: (string filename)
output: eight lists containing all data entries for all eight columns

Description: 
Given the path for an excel spreadsheet,
go through each row and add the elements from each column to a corresponding list.

Note: This is the format of the columns from excel sheet
['Subject', 'Course #', 'Course Title', 'Ver.', 'Sec.', 'Instructor Real Name', 'Time', 'Capacity']
"""

def load_schedule_hardcoded(filename):
	#Either of these work, it's a matter of preference
	subjects, course_nums, course_titles, versions, sections, instructor_real_names, times, capacities = [], [], [], [], [], [], [], []

	"""
	subjects = []
	course_nums = []
	course_titles = []
	versions = []
	sections = []
	instructor_real_names = []
	times = []
	capacities = []
	"""
	
	try:
		workbook = load_workbook(filename) 		#Load .xlsx file
	except:
	    error_file_not_found(filename)
	
	sheet = workbook.active				#Current excel sheet (we are only using one)
	rows = sheet.rows				#Rows from excel sheet
	max_row = sheet.max_row				#Gets count of max row
	
	"""
	Why start the loop at 2:
		We do not need the first row which contains the headers i.e. subject, course #, etc.
		Additionally, iterating through an excel spreadsheet starts at 1 not 0.		
		
	Why end the loop at max_row + 1:
		Since the iteration starts at one, we need to add +1 at the end to get every row
	"""
	for i in range(2, max_row + 1): #do not need first row
		
		"""
		According to Dutt the excel sheet we will be given will always be the same format
		This is hardcoded based on that format, but it is easy to change as needed.
		"""
		subject = sheet.cell(row=i, column=1)
		subjects.append(subject.value)
		
		course_num = sheet.cell(row=i, column=2) 
		course_nums.append(str(course_num.value)) #force str because of elems like 201ss
		
		course_title = sheet.cell(row=i, column=3)
		course_titles.append(course_title.value)

		version = sheet.cell(row=i, column=4)
		versions.append(version.value)

		section = sheet.cell(row=i, column=5)
		sections.append(section.value)

		instructor_real_name = sheet.cell(row=i, column=6)
		instructor_real_names.append(instructor_real_name.value)

		time = sheet.cell(row=i, column=7)
		times.append(time.value)

		capacity = sheet.cell(row=i, column=8)
		capacities.append(capacity.value)
			
	return subjects, course_nums, course_titles, versions, sections, instructor_real_names, times, capacities


"""
load classrooms hardcoded

input: (string filename)
output: two lists containing all data entries for all two columns

Description: 
Given the path for an excel spreadsheet,
go through each row and add the elements from each column to a corresponding list.

Note: This is the format of the columns from excel sheet
['Classroom', 'Capacity']
"""

def load_classrooms_hardcoded(filename):       
	#Either of these work, it's a matter of preference
	classrooms, capacities = [], []
	"""
	classroom = []
	capacity = []
	"""
	
	try:
	    workbook = load_workbook(filename) 		#Load .xlsx file
	except:
		error_file_not_found(filename)
	
	sheet = workbook.active				#Current excel sheet (we are only using one)
	rows = sheet.rows				#Rows from excel sheet
	max_row = sheet.max_row				#Gets count of max row
	
	"""
	Why start the loop at 2:
		We do not need the first row which contains the headers i.e. subject, course #, etc.
		Additionally, iterating through an excel spreadsheet starts at 1 not 0.		
		
	Why end the loop at max_row + 1:
		Since the iteration starts at one, we need to add +1 at the end to get every row
	"""
	for i in range(2, max_row + 1): #do not need first row
		
		"""
		According to Dutt the excel sheet we will be given will always be the same format
		This is hardcoded based on that format, but it is easy to change as needed.
		"""
		classroom = sheet.cell(row=i, column=1)
		classrooms.append(classroom.value)
		
		capacity = sheet.cell(row=i, column=2) 
		capacities.append(capacity.value)
	
	return classrooms, capacities


